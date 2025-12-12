import re
import os
import openpyxl
from openpyxl.utils import get_column_letter
from concurrent.futures import ThreadPoolExecutor
import glob


def read_data_from_txt(txt_path):
    """从txt文件中读取指定行的数据"""
    scalar_lines = list(range(6, 172, 15))  # 6, 21, ..., 171
    array_lines = list(range(8, 174, 15))  # 8, 23, ..., 173

    scalars = {}
    arrays = {}

    try:
        with open(txt_path, 'r', encoding='utf-8') as file:
            for i, line in enumerate(file, 1):
                if i in scalar_lines:
                    value = float(line.strip())
                    scalars[i] = value
                elif i in array_lines:
                    line = line.strip()
                    array_str = re.search(r'\[([^\]]+)\]', line).group(1)
                    array = [float(num) for num in re.findall(r'[-+]?\d*\.\d+|\d+', array_str)]
                    arrays[i] = array
    except UnicodeDecodeError:
        try:
            with open(txt_path, 'r', encoding='gb18030') as file:
                for i, line in enumerate(file, 1):
                    if i in scalar_lines:
                        value = float(line.strip())
                        scalars[i] = value
                    elif i in array_lines:
                        line = line.strip()
                        array_str = re.search(r'\[([^\]]+)\]', line).group(1)
                        array = [float(num) for num in re.findall(r'[-+]?\d*\.\d+|\d+', array_str)]
                        arrays[i] = array
        except Exception as e:
            print(f"无法读取文件 {txt_path}: {e}")
            return {}, {}

    return scalars, arrays


def write_to_excel(excel_path, all_data, template_path):
    """将多个文件的数据写入Excel文件"""
    try:
        # 复制模板文件
        if os.path.exists(template_path):
            workbook = openpyxl.load_workbook(template_path)
        else:
            # 如果模板不存在，创建新工作簿
            workbook = openpyxl.Workbook()

        sheet = workbook.active

        # 定义每个算法的起始列
        algorithm_start_cols = {
            'ISO': 2,  # B列
            'SO': 14,  # N列
            'MWOA': 26,  # Z列
            'JCGA': 38,  # AL列
            'IABC': 50  # AX列
        }

        # 写入所有算法的数据
        for algorithm, (scalars, arrays) in all_data.items():
            start_col = algorithm_start_cols[algorithm]

            # 写入标量数据到第3行
            for line_num, value in scalars.items():
                col_offset = (line_num - 6) // 15
                col_letter = get_column_letter(start_col + col_offset)
                sheet[f'{col_letter}3'] = value

            # 写入数组数据
            for line_num, array in arrays.items():
                col_offset = (line_num - 8) // 15
                col_letter = get_column_letter(start_col + col_offset)

                for row_idx, value in enumerate(array, start=4):
                    sheet[f'{col_letter}{row_idx}'] = value

        # 保存工作簿
        workbook.save(excel_path)
        print(f"成功生成Excel文件: {excel_path}")
    except Exception as e:
        print(f"生成Excel文件失败 {excel_path}: {e}")


def process_txt_group(txt_group, template_path, output_dir):
    """处理一组同名txt文件并生成Excel"""
    if not txt_group:
        return

    # 从第一个txt文件路径提取信息
    first_path = txt_group[0][1]
    # 提取规模 (small, middle, big)
    scale_match = re.search(r'(small|middle|big)\\SET\d+\\', first_path)
    scale = scale_match.group(1) if scale_match else "unknown"

    # 提取SET编号
    set_match = re.search(r'(SET\d+)\\', first_path)
    set_number = set_match.group(1) if set_match else "SET0"

    # 提取L和P参数
    param_match = re.search(r'Instance_L\((.+?)\)_P\((.+?)\)_', first_path)
    l_params = param_match.group(1) if param_match else "x,x,x"
    p_params = param_match.group(2) if param_match else "x,x,x"

    # 构建输出文件名
    output_filename = f"Fit_list_{scale}_{set_number}_L({l_params})_P({p_params}).xlsx"
    output_path = os.path.join(output_dir, output_filename)

    # 读取所有算法的数据
    all_data = {}
    for algorithm, txt_path in txt_group:
        scalars, arrays = read_data_from_txt(txt_path)
        all_data[algorithm] = (scalars, arrays)

    # 写入Excel
    write_to_excel(output_path, all_data, template_path)


def main():
    # 定义算法列表
    algorithms = ['ISO', 'SO', 'MWOA', 'JCGA', 'IABC']

    # 定义规模列表
    scales = ['small', 'middle', 'big']

    # 定义SET列表
    sets = [f'SET{i}' for i in range(1, 6)]

    # 模板文件路径
    template_path = 'Fit_list.xlsx'

    # 输出目录
    output_dir = 'output_excels'
    os.makedirs(output_dir, exist_ok=True)

    # 构建所有可能的txt文件路径模式
    txt_patterns = []
    for algorithm in algorithms:
        for scale in scales:
            for set_num in sets:
                pattern = os.path.join(
                    'result3',
                    f'{algorithm}_result3',
                    'randomDatasetResult',
                    scale,
                    set_num,
                    'Instance_L(*)_P(*)_*_result.txt'
                )
                txt_patterns.append((algorithm, pattern))

    # 按文件名分组txt文件
    txt_groups = {}
    for algorithm, pattern in txt_patterns:
        for txt_path in glob.glob(pattern):
            # 提取文件名部分（不含扩展名）
            filename = os.path.basename(txt_path)
            base_name = os.path.splitext(filename)[0]
            # 移除算法名称部分，使不同算法的同名文件能匹配
            key_name = re.sub(r'_' + algorithm + '_result', '_*_result', base_name)

            if key_name not in txt_groups:
                txt_groups[key_name] = []
            txt_groups[key_name].append((algorithm, txt_path))

    # 处理每个文件组
    with ThreadPoolExecutor(max_workers=5) as executor:
        for key_name, txt_group in txt_groups.items():
            executor.submit(process_txt_group, txt_group, template_path, output_dir)

    print(f"全部处理完成！共生成 {len(txt_groups)} 个Excel文件。")


if __name__ == "__main__":
    main()