import os
import re
from openpyxl import load_workbook
#按文件名处理SET文件夹中的txt文件，做到了正确匹配，并且SOA算法按照顺序替换从12个选出10个较小值

def main():
    # 设置文件夹路径和Excel文件路径
    base_folder = r'result\ISO_result_redo\randomDatasetResult\big'
    excel_file = 'temp.xlsx'

    # 加载Excel文件
    try:
        wb = load_workbook(excel_file)
        ws = wb['big']
    except Exception as e:
        print(f"加载Excel文件时出错: {e}")
        return

    # 创建一个字典，用于存储每个SET文件夹中的txt文件，键为L和P值的元组
    txt_files_dict = {}
    for set_num in range(1, 6):
        set_folder = os.path.join(base_folder, f'SET{set_num}')
        if not os.path.exists(set_folder):
            print(f"SET文件夹不存在: {set_folder}")
            continue

        # 获取该SET文件夹中的所有txt文件
        txt_files = [f for f in os.listdir(set_folder) if f.endswith('.txt')]

        # 处理每个txt文件，提取L和P值
        for txt_file in txt_files:
            match = re.search(r'Instance_L\((\d+,\d+,\d+)\)_P\((\d+,\d+,\d+)\)_ISO_result\.txt', txt_file)
            if match:
                l_values = match.group(1)
                p_values = match.group(2)
                key = f"L({l_values})*P({p_values})"
                txt_files_dict[key] = os.path.join(set_folder, txt_file)

    # 处理Excel中的合并单元格
    for i in range(0, 40):
        # 计算合并单元格的范围和数据单元格的范围
        start_row = 3 + i * 15
        merge_range = f"B{start_row}:B{start_row + 14}"

        # 检查该合并单元格是否存在
        is_merged = False
        for merged_range in ws.merged_cells.ranges:
            if str(merged_range) == merge_range:
                is_merged = True
                # 获取合并单元格的起始单元格
                min_col, min_row, max_col, max_row = merged_range.bounds
                cell_value = ws.cell(row=min_row, column=min_col).value
                break

        if not is_merged:
            print(f"合并单元格 {merge_range} 不存在")
            continue

        # 检查单元格值是否符合格式
        if not cell_value or not isinstance(cell_value, str):
            print(f"合并单元格 {merge_range} 的值为空或不是字符串")
            continue

        # 提取L和P值
        match = re.search(r'L\((\d+,\d+,\d+)\)\*P\((\d+,\d+,\d+)\)', cell_value)
        if not match:
            print(f"合并单元格 {merge_range} 的值格式不正确: {cell_value}")
            continue

        # 构建查找键
        key = cell_value

        # 查找对应的txt文件
        if key in txt_files_dict:
            txt_file_path = txt_files_dict[key]
            try:
                # 使用UTF-8编码打开文件，如果失败尝试latin-1
                try:
                    with open(txt_file_path, 'r', encoding='utf-8') as f:
                        lines = f.readlines()
                except UnicodeDecodeError:
                    with open(txt_file_path, 'r', encoding='latin-1') as f:
                        lines = f.readlines()

                # 提取所需行的数据（第6行、第21行、...、第141行，以及第156行和第171行）
                data = []
                # 读取前10个值
                for i_line in range(6, 142, 15):
                    if i_line <= len(lines):
                        try:
                            value = float(lines[i_line - 1].strip())
                            data.append(value)
                        except ValueError:
                            print(f"无法将第{i_line}行转换为数值: {lines[i_line - 1].strip()}")
                            data.append(None)
                    else:
                        print(f"文件 {txt_file_path} 没有第{i_line}行")
                        data.append(None)

                # 读取第11个和第12个值
                extra_values = []
                for i_line in [156, 171]:
                    if i_line <= len(lines):
                        try:
                            value = float(lines[i_line - 1].strip())
                            extra_values.append(value)
                        except ValueError:
                            print(f"无法将第{i_line}行转换为数值: {lines[i_line - 1].strip()}")
                            extra_values.append(None)
                    else:
                        print(f"文件 {txt_file_path} 没有第{i_line}行")
                        extra_values.append(None)

                # 确保extra_values有两个有效的值
                valid_extra_values = [v for v in extra_values if v is not None]
                if len(valid_extra_values) == 2:
                    a, b = sorted(valid_extra_values)  # a、b是第11个和第12个值，其中a是较小值，b是较大值

                    # 获取原始10个值中的最小值
                    original_min = min(v for v in data if v is not None)

                    # 执行替换规则
                    if b > original_min:
                        # 找到最小值的位置并替换为b
                        min_index = data.index(original_min)
                        data[min_index] = b

                        # 替换后，获取新的最小值
                        new_min = min(v for v in data if v is not None)

                        # 检查a是否大于新的最小值
                        if a > new_min:
                            # 找到新的最小值位置并替换为a
                            new_min_index = data.index(new_min)
                            data[new_min_index] = a

                    print(f"应用替换规则: a={a}, b={b}, 原始最小值={original_min}")

                # 计算在Excel中的起始单元格
                data_start_row = 4 + i * 15
############################################不同算法改这里
                # 将数据写入Excel
                for j, value in enumerate(data):
                    cell = f'D{data_start_row + j}'
                    ws[cell] = value

                print(f"已处理: {cell_value} -> D{data_start_row}:D{data_start_row + 9}")
############################################
            except Exception as e:
                print(f"处理文件 {txt_file_path} 时出错: {e}")
        else:
            print(f"找不到对应的txt文件: {cell_value}")

    # 保存Excel文件
    try:
        wb.save(excel_file)
        print("Excel文件已保存")
    except Exception as e:
        print(f"保存Excel文件时出错: {e}")


if __name__ == "__main__":
    main()