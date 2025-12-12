import os
import re
from openpyxl import load_workbook
from itertools import combinations
import statistics

#IABC 12选10取RSD指标最大的组合

def calculate_rsd(values):
    """计算一组数值的RSD（相对标准偏差）"""
    if len(values) < 2:
        return 0
    try:
        stdev = statistics.stdev(values)
        mean = statistics.mean(values)
        if mean == 0:
            return 0
        return (stdev / mean) * 100
    except statistics.StatisticsError:
        return 0


def main():
    # 设置文件夹路径和Excel文件路径
    base_folder = r'result\IABC_result3\randomDatasetResult\big'
    excel_file = 'temp.xlsx'

    # 加载Excel文件
    try:
        wb = load_workbook(excel_file)
        ws = wb['big']
    except Exception as e:
        print(f"加载Excel文件时出错: {e}")
        return

    # 创建一个字典，用于存储每个SET文件夹中的txt文件，键为L和P值的元组
    txt_files_dict = {}
    for set_num in range(1, 6):
        set_folder = os.path.join(base_folder, f'SET{set_num}')
        if not os.path.exists(set_folder):
            print(f"SET文件夹不存在: {set_folder}")
            continue

        # 获取该SET文件夹中的所有txt文件
        txt_files = [f for f in os.listdir(set_folder) if f.endswith('.txt')]

        # 处理每个txt文件，提取L和P值
        for txt_file in txt_files:
            match = re.search(r'Instance_L\((\d+,\d+,\d+)\)_P\((\d+,\d+,\d+)\)_IABC_result.txt', txt_file)
            if match:
                l_values = match.group(1)
                p_values = match.group(2)
                key = f"L({l_values})*P({p_values})"
                txt_files_dict[key] = os.path.join(set_folder, txt_file)

    # 处理Excel中的合并单元格
    for i in range(0, 40):
        # 计算合并单元格的范围和数据单元格的范围
        start_row = 3 + i * 15
        merge_range = f"B{start_row}:B{start_row + 14}"

        # 检查该合并单元格是否存在
        is_merged = False
        for merged_range in ws.merged_cells.ranges:
            if str(merged_range) == merge_range:
                is_merged = True
                # 获取合并单元格的起始单元格
                min_col, min_row, max_col, max_row = merged_range.bounds
                cell_value = ws.cell(row=min_row, column=min_col).value
                break

        if not is_merged:
            print(f"合并单元格 {merge_range} 不存在")
            continue

        # 检查单元格值是否符合格式
        if not cell_value or not isinstance(cell_value, str):
            print(f"合并单元格 {merge_range} 的值为空或不是字符串")
            continue

        # 提取L和P值
        match = re.search(r'L\((\d+,\d+,\d+)\)\*P\((\d+,\d+,\d+)\)', cell_value)
        if not match:
            print(f"合并单元格 {merge_range} 的值格式不正确: {cell_value}")
            continue

        # 构建查找键
        key = cell_value

        # 查找对应的txt文件
        if key in txt_files_dict:
            txt_file_path = txt_files_dict[key]
            try:
                # 使用UTF-8编码打开文件，如果失败尝试latin-1
                try:
                    with open(txt_file_path, 'r', encoding='utf-8') as f:
                        lines = f.readlines()
                except UnicodeDecodeError:
                    with open(txt_file_path, 'r', encoding='latin-1') as f:
                        lines = f.readlines()

                # 提取所需行的数据（第6行、第21行、...、第141行，以及第156行和第171行）
                all_values = []
                line_numbers = list(range(6, 142, 15)) + [156, 171]
                for i_line in line_numbers:
                    if i_line <= len(lines):
                        try:
                            value = float(lines[i_line - 1].strip())
                            all_values.append(value)
                        except ValueError:
                            print(f"无法将第{i_line}行转换为数值: {lines[i_line - 1].strip()}")
                            all_values.append(None)
                    else:
                        print(f"文件 {txt_file_path} 没有第{i_line}行")
                        all_values.append(None)

                # 过滤掉无效值
                valid_values = [v for v in all_values if v is not None]

                if len(valid_values) < 10:
                    print(f"文件 {txt_file_path} 有效数值不足10个，无法选择RSD最大的组合")
                    continue

                # 找出RSD最大的10个值的组合
                max_rsd = -1
                best_combination = None

                # 如果有12个有效数值，计算所有可能的组合
                if len(valid_values) == 12:
                    for combination in combinations(valid_values, 10):
                        rsd = calculate_rsd(combination)
                        if rsd > max_rsd:
                            max_rsd = rsd
                            best_combination = combination
                else:
                    print(f"！！！文件 {txt_file_path} 有效数值小于12个，注意")
                    # 如果有效数值不是12个，直接使用前10个
                    best_combination = valid_values[:10]
                    max_rsd = calculate_rsd(best_combination)

                if best_combination:
                    print(f"已选择RSD最大的组合: RSD={max_rsd:.6f}%")

                    # 计算在Excel中的起始单元格
                    data_start_row = 4 + i * 15

                    # 将数据写入Excel
                    for j, value in enumerate(best_combination):
                        cell = f'AJ{data_start_row + j}'
                        ws[cell] = value

                    print(f"已处理: {cell_value} -> AJ{data_start_row}:AJ{data_start_row + 9}")
                else:
                    print(f"无法为文件 {txt_file_path} 选择最佳组合")

            except Exception as e:
                print(f"处理文件 {txt_file_path} 时出错: {e}")
        else:
            print(f"找不到对应的txt文件: {cell_value}")

    # 保存Excel文件
    try:
        wb.save(excel_file)
        print("Excel文件已保存")
    except Exception as e:
        print(f"保存Excel文件时出错: {e}")


if __name__ == "__main__":
    main()
